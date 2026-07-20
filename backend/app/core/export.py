"""Xuất BOQ/khối lượng cáp & ống từ dữ liệu đã bóc & duyệt.

Nguồn: review_store (dữ liệu người dùng đã sửa/xác nhận theo trang). Hỗ trợ:
  - JSON gộp toàn dự án.
  - Excel 2 sheet: 'Chi tiết' (từng lộ) + 'Tổng hợp' (đếm theo tiết diện/quy cách).
"""
from __future__ import annotations

import io
import json
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..storage import review_store
from .takeoff.sheet_meta import normalize_meta

# (khoá field, tiêu đề cột, độ rộng)
# 5 cột đầu là thông tin hành chính của TRANG (từ rec['meta']), được gắn vào mọi
# dòng của trang đó để mỗi dòng BOQ tự đứng độc lập khi lọc/gộp trong Excel.
COLUMNS = [
    ("page", "Trang", 8),
    ("project", "Dự án / Hạng mục", 24),
    ("towerOrBasement", "Tháp / Hầm", 14),
    ("area", "Khu vực", 22),
    ("floor", "Tầng", 10),
    ("typicalCode", "Điển hình", 18),
    ("panelName", "Tủ / Khu vực", 30),
    ("roadName", "Lộ", 10),
    ("loadName", "Tên tải", 30),
    ("itemGroup", "Nhóm vật tư", 16),
    ("itemName", "Tên vật tư", 18),
    ("size", "Tiết diện", 12),
    ("cableSpec", "Quy cách cáp", 42),
    ("conduit", "Ống luồn", 22),
]


def collect_items(job_id: str, only_confirmed: bool = False) -> list[dict]:
    data = review_store.get_all(job_id)
    rows: list[dict] = []
    for _, rec in sorted(data["pages"].items(), key=lambda kv: int(kv[0])):
        if only_confirmed and not rec.get("confirmed"):
            continue
        meta = normalize_meta(rec.get("meta"))
        for it in rec.get("items", []):
            row = dict(it)
            row["page"] = rec["page"]
            # Thông tin trang đứng TRƯỚC field của item: item không có khoá nào
            # trùng tên nên không ghi đè, nhưng đặt vậy cho rõ thứ tự ưu tiên.
            row.update(meta)
            rows.append(row)
    return rows


def collect_panels(job_id: str, only_confirmed: bool = False) -> list[dict]:
    data = review_store.get_all(job_id)
    out, seen = [], set()
    for _, rec in sorted(data["pages"].items(), key=lambda kv: int(kv[0])):
        if only_confirmed and not rec.get("confirmed"):
            continue
        for p in rec.get("panels", []):
            key = p.get("name", "")
            if key and key not in seen:
                seen.add(key)
                out.append({**p, "page": rec["page"]})
    return out


def to_json(job_id: str, only_confirmed: bool = False) -> list[dict]:
    """Danh sách item theo schema QS (giữ nguyên các field gốc)."""
    items = collect_items(job_id, only_confirmed)
    for it in items:
        it.pop("page", None)  # JSON QS không cần page
    return items


def to_json_text(job_id: str, only_confirmed: bool = False) -> str:
    """JSON dễ đọc: thụt lề chuẩn (indent=2), tiếng Việt giữ nguyên."""
    items = to_json(job_id, only_confirmed)
    return json.dumps(items, ensure_ascii=False, indent=2)


def to_excel(job_id: str, only_confirmed: bool = False) -> bytes:
    items = collect_items(job_id, only_confirmed)
    wb = Workbook()

    # --- Sheet chi tiết ---
    ws = wb.active
    ws.title = "Chi tiết"
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["STT"] + [c[1] for c in COLUMNS]
    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, it in enumerate(items, start=1):
        ws.append([idx] + [it.get(k, "") for k, _, _ in COLUMNS])

    ws.column_dimensions["A"].width = 6
    for i, (_, _, w) in enumerate(COLUMNS, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Sheet tổng hợp (đếm số lộ theo tiết diện/quy cách/ống) ---
    ws2 = wb.create_sheet("Tổng hợp")
    agg: dict[tuple, int] = defaultdict(int)
    for it in items:
        key = (it.get("size", ""), it.get("cableSpec", ""), it.get("conduit", ""))
        if any(key):
            agg[key] += 1
    sum_headers = ["STT", "Tiết diện", "Quy cách cáp", "Ống luồn", "Số lộ"]
    ws2.append(sum_headers)
    for col in range(1, len(sum_headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, (key, n) in enumerate(sorted(agg.items()), start=1):
        ws2.append([idx, key[0], key[1], key[2], n])
    for col, w in zip("ABCDE", (6, 14, 44, 24, 10)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet tủ tổng (nếu có) ---
    panels = collect_panels(job_id, only_confirmed)
    if panels:
        ws3 = wb.create_sheet("Tủ điện")
        ph = ["STT", "Tủ điện", "Loại", "Công suất P (W)", "Ptt (W)", "Kđt"]
        ws3.append(ph)
        for col in range(1, len(ph) + 1):
            c = ws3.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        kind_label = {"main": "Tủ tổng", "sub": "Tủ con (smarthome)"}
        for i, p in enumerate(panels, start=1):
            ws3.append([i, p.get("name", ""), kind_label.get(p.get("kind"), ""),
                        p.get("power", ""), p.get("ptt", ""), p.get("kdt", "")])
        for col, w in zip("ABCDEF", (6, 16, 18, 16, 14, 10)):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
